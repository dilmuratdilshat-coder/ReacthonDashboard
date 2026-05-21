from openpyxl import load_workbook
import csv
import os
#from itertools import 

# Открываем файл
file_path = "data/ВОДА (Автосохраненный) 1.xlsx"
wb = load_workbook(file_path, data_only=True)

# Определяем заголовки для обеих таблиц
headers_config = [
    ['Дата', 'Номинальный ежедневный расход', 'показание счетчика',
     'Фактический расход за сутки', 'Расход в час'],
    ['Дата', 'показание счетчика', 'Расход за сутки', 'Расход в час']
]

all_tables = []

def has_error(cell_value):
    """Проверяет, содержит ли ячейка ошибку Excel"""
    if cell_value is None:
        return False
    cell_str = str(cell_value).strip()
    # Проверяем на типичные ошибки Excel
    error_indicators = ['#ССЫЛКА!', '#REF!', '#VALUE!', '#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!']
    return any(error in cell_str for error in error_indicators)

def row_has_error(row_data):
    """Проверяет, есть ли ошибка в строке данных"""
    return any(has_error(cell) for cell in row_data)

# Обрабатываем каждый лист
for sheet_name in wb.sheetnames:
    print(f"Поиск в листе: {sheet_name}")
    sheet = wb[sheet_name]

    rows_list = list(sheet.iter_rows(values_only=True))

    for i, row in enumerate(rows_list):
        row_list = list(row)

        # Проверяем каждый вариант заголовка
        for target_header in headers_config:
            # Ищем последовательность заголовков в любом месте строки
            for start_col in range(len(row_list) - len(target_header) + 1):
                # Проверяем совпадение с этой позиции
                if row_list[start_col:start_col + len(target_header)] == target_header:

                    # Добавляем заголовок (только нужные колонки)
                    table_data = [row_list[start_col:start_col + len(target_header)]]

                    # Читаем данные до первой пустой строки
                    j = i + 1
                    skipped_rows = 0

                    while j < len(rows_list):
                        data_row = list(rows_list[j])

                        # Берем те же колонки
                        data_cols = data_row[start_col:start_col + len(target_header)]

                        # Проверяем, пустая ли строка в этих колонках
                        if all(cell is None or str(cell).strip() == '' for cell in data_cols):
                            break

                        # Проверяем на ошибки Excel
                        if row_has_error(data_cols):
                            skipped_rows += 1
                            j += 1
                            continue

                        table_data.append(data_cols)
                        j += 1

                    # Добавляем разделитель между таблицами
                    if all_tables:
                        all_tables.append([])

                    all_tables.extend(table_data)
                    #print(*all_tables, sep='\n\n')


if __name__ == '__main__':
    # Сохраняем все таблицы в один файл
    os.makedirs('output', exist_ok=True)
    output_file = "output/extracted_tables_clean.csv"

    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter='\t')
        def nom_tester(lst):
            res = lst if (len(lst) == 5) else [lst[0], float('NaN'), *lst[1:]] 
            #print(res, lst, type(lst))
            return res
        all_tables = [all_tables[0]] + [nom_tester(lst) for lst in all_tables[1:] if not (None in lst or lst == [] or isinstance(lst[0], str))]
        writer.writerows(all_tables)


    wb.close()